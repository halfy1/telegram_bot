from openpyxl import load_workbook
import matplotlib.pyplot as plt
import numpy as np

def main(option, message, bot):
    if option == "Калибровка установки":
        return opt1(message, bot)
    elif option == "Вывод":
        return opt2(message, bot)
    elif option == "Расчёт":
        return opt3(message, bot)
    elif option == "Графики":
        return opt4(message, bot)
    elif option == "Созвон":
        return opt5(message, bot)

result1 = f"В ходе данной работы изучалось распределение молекул идеального газа по координатам (по высоте) при помощи его механической модели. Были построены графики зависимостей логарифма количества пересечений стеклянных шариков (моделей молекул идеального газа) луча фотодатчика от высоты для двух частот колебаний основания установки, что имитировало различные температуры газа. Эти зависимости носят линейный характер, следовательно, можно считать экспериментально подтвержденной теоретически выведенную зависимость ln N = const – mg/kT*h. \
Методом парных точек были найдены угловые коэффициенты, отношение которых можно сравнить с отношением температур (т.е. частот).\n"
result2= f"α1 = (…±…)мм   ε = …%     α = 0,9  \nα2 = (…±…)мм   ε = …%     α = 0,9\nηэксп = (…±…)  ε = …% \nηтеор = …\n "
result3=f"Теоретическое значение отношения входит в интервал экспериментального значения отношения с учетом погрешностей, что подтверждает теоретическое распределение молекул газа по высоте в поле сил тяжести.\n"
result=result1+result2+result3
def opt1(message, bot):
    bot.send_message(message.chat.id, "Вы выбрали 'Оптимальные параметры установки'")
    bot.send_message(message.chat.id, "Вам всё расскажут во время лабораторной работы) Остальное читать в лабнике")
def opt2(message, bot):
    bot.send_message(message.chat.id, "Вы выбрали 'Составление вывода'")
    bot.send_message(message.chat.id, result)

USER_STATE = {}

def get_user_state(user_id):
    return USER_STATE.get(user_id, 'NORMAL')

def update_user_state(user_id, state):
    USER_STATE[user_id] = state

def opt3(message, bot):
    bot.send_message(message.chat.id, "Пожалуйста, введите 16 строк, каждая из которых содержит 4 числа, разделенных пробелами. Если вы хотите остановиться, введите /stop")
    return 'INPUT_FOR_OPT3'

def opt4(message, bot):
    bot.send_message(message.chat.id, "Пожалуйста, введите 8 строк, каждая из которых содержит 2 числа, разделенных пробелами. Если вы хотите остановиться, введите /stop")
    return 'INPUT_FOR_OPT4'

def handle_input_for_opt4(message, bot):
    if message.text == "/stop":
        bot.send_message(message.chat.id, "Остановлено пользователем")
        return 'NORMAL'

    lines = message.text.split("\n")
    if len(lines) != 8:
        bot.send_message(message.chat.id, "Пожалуйста, введите ровно 8 строк. Попробуйте еще раз.")
        return 'INPUT_FOR_OPT4'

    h = []
    lnN = []
    for line in lines:
        numbers = line.split()
        if len(numbers) != 2:
            bot.send_message(message.chat.id, "Каждая строка должна содержать ровно 2 числа. Попробуйте еще раз.")
            return 'INPUT_FOR_OPT4'
        try:
            h.append(float(numbers[0]))
            lnN.append(float(numbers[1]))
        except ValueError:
            bot.send_message(message.chat.id, "Пожалуйста, убедитесь, что вы вводите только числа. Попробуйте еще раз.")
            return 'INPUT_FOR_OPT4'

    # Строим график
    fig, ax = plt.subplots(figsize=(11.69, 8.27))  # Размер листа А4 в дюймах

    # Отображаем отдельные точки
    ax.scatter(h, lnN, label='ln(N)', color='red')

    # Выполняем линейную аппроксимацию
    coefficients = np.polyfit(h, lnN, 1)
    polynomial = np.poly1d(coefficients)
    hs = np.linspace(min(h), max(h), 100)
    ax.plot(hs, polynomial(hs), label='Аппроксимация')

    ax.set_title('Зависимость ln(N) от h')
    ax.set_xlabel('h')
    ax.set_ylabel('ln(N)')
    ax.legend()

    # Добавляем сетку
    ax.grid(True)

    # Сохраняем график в файл PDF
    plt.savefig('graph.pdf')

    # Отправляем файл пользователю
    with open('graph.pdf', 'rb') as graph:
        bot.send_document(message.chat.id, graph)

    bot.send_message(message.chat.id, "График успешно построен и сохранен в файл 'graph.pdf'. Файл был отправлен вам.")

    return 'NORMAL'


def opt5(message, bot):
    bot.send_message(message.chat.id, "Вы выбрали 'Личный созвон'")

from openpyxl import load_workbook

def handle_input_for_opt3(message, bot):
    if message.text == "/stop":
        bot.send_message(message.chat.id, "Остановлено пользователем")
        return 'NORMAL'

    lines = message.text.split("\n")
    if len(lines) != 16:
        bot.send_message(message.chat.id, "Пожалуйста, введите ровно 16 строк. Попробуйте еще раз.")
        return 'INPUT_FOR_OPT3'

    h = []
    N1 = []
    N2 = []
    N3 = []
    for line in lines:
        numbers = line.split()
        if len(numbers) != 4:
            bot.send_message(message.chat.id, "Каждая строка должна содержать ровно 4 числа. Попробуйте еще раз.")
            return 'INPUT_FOR_OPT3'
        try:
            h.append(float(numbers[0]))
            N1.append(float(numbers[1]))
            N2.append(float(numbers[2]))
            N3.append(float(numbers[3]))
        except ValueError:
            bot.send_message(message.chat.id, "Пожалуйста, убедитесь, что вы вводите только числа. Попробуйте еще раз.")
            return 'INPUT_FOR_OPT3'

    # Загружаем существующий файл Excel
    workbook = load_workbook(filename="2.23.xlsx")
    sheet = workbook.active

    # Записываем первые 8 значений в ячейки E7:H14
    for i in range(8):
        sheet[f"E{i+7}"] = h[i]
        sheet[f"F{i+7}"] = N1[i]
        sheet[f"G{i+7}"] = N2[i]
        sheet[f"H{i+7}"] = N3[i]

    # Записываем оставшиеся 8 значений в ячейки E17:H24
    for i in range(8, 16):
        sheet[f"E{i+9}"] = h[i]
        sheet[f"F{i+9}"] = N1[i]
        sheet[f"G{i+9}"] = N2[i]
        sheet[f"H{i+9}"] = N3[i]

    # Сохраняем изменения
    workbook.save(filename="2.23.xlsx")

    bot.send_message(message.chat.id, "Ваши данные успешно обработаны и записаны в файл Excel!")

    bot.send_document(message.chat.id, open('2.23.xlsx', 'rb'))

    return 'NORMAL'
